"""

pre-request:

    pip install Bio
    pip install openpyxl

//*[@id="summaryDl"]/dd[10]

demo_url = "https://www.ncbi.nlm.nih.gov/gene/7422"
full_url = "https://www.ncbi.nlm.nih.gov/gene/7422?report=full_report&format=text"


KLF2

Created at 2023/9/14
"""

import openpyxl
from Bio import Entrez

from openpyxl.styles import Border, Font, PatternFill, Side

Entrez.email = '1052433260@qq.com'

data_path = r'C:\Users\libyao\Downloads\genes-LH.xlsx'
output_filename = 'output.xlsx'


def get_gene_id(gene_symbol: str):
    gene_term = f"({gene_symbol}[Gene Name]) AND Homo sapiens[Organism]"
    try:
        handle = Entrez.esearch(db="gene", term=gene_term)
        gene_id = Entrez.read(handle)['IdList']
    except:
        return None

    if not gene_id:
        return None

    # 10365
    # print(gene_id)
    return gene_id[0]


def get_summary_info(gene_id):
    res = dict()
    try:
        sum_handle = Entrez.esummary(db="gene", id=gene_id)
        sum_record = Entrez.read(sum_handle)
    except:
        return {}
    # pprint.pprint(sum_record)

    data = sum_record['DocumentSummarySet']['DocumentSummary']
    if not data:
        return {}

    data = data[0]
    res['desc'] = data['Description']
    res['summary'] = data['Summary']

    # pprint.pprint(res)
    return res


def read_excel_gene_symbol(file_path):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path)

    # 选择要读取数据的工作表（这里假设要读取第一个工作表）
    sheet = workbook.active

    # 通过遍历每一行来获取第一列的数据
    first_column_data = []

    # min_row = 2 不读取表头
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        first_column_data.append(row[0])

    # # 打印第一列的数据
    # for data in first_column_data:
    #     print(data)

    return first_column_data


def main_process():
    result_list = []
    missed_list = []

    count = 0

    result_list.append(['Gene Symbal', 'Url', 'Desc', 'Summary'])

    gene_symbol_list = read_excel_gene_symbol(data_path)

    print(f'start to download [{len(gene_symbol_list)}] Gene data')

    for symbol in gene_symbol_list:
        gene_id = get_gene_id(symbol)
        if not gene_id:
            missed_list.append(symbol)
            continue

        gene_url = f'https://www.ncbi.nlm.nih.gov/gene/{gene_id}'

        res = get_summary_info(gene_id)
        print(f'{symbol}', end=' ')
        count += 1
        if count % 20 == 0:
            print()

        result_list.append([symbol, gene_url, res.get('desc', ''), res.get('summary', '')])

    print()
    print(f'Missed symbol: {missed_list}')
    save_to_excel(output_filename, result_list)


def save_to_excel(filename, final_datas):
    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()

    # 获取默认的工作表Sheet
    sheet = wb.active

    # 向Sheet添加表头
    headers = final_datas[0]
    for i, header in enumerate(headers):
        sheet.cell(row=1, column=i + 1, value=header)

    # 添加表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='BDD7EE', end_color='BDD7EE')
    header_border = Border(bottom=Side(border_style='thin'))
    for i in range(1, len(headers) + 1):
        sheet.cell(row=1, column=i).font = header_font
        sheet.cell(row=1, column=i).fill = header_fill
        sheet.cell(row=1, column=i).border = header_border

    # 向Sheet添加数据
    data = final_datas[1:]

    for r, row in enumerate(data, start=2):
        for c, cell_value in enumerate(row, start=1):
            sheet.cell(row=r, column=c, value=cell_value)

    # 添加数据行样式
    data_border = Border(bottom=Side(border_style='thin', color='D9D9D9'))
    for r in range(2, len(data) + 2):
        for c in range(1, len(headers) + 1):
            sheet.cell(row=r, column=c).border = data_border

    # 自适应列宽
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # 保存Excel文件
    wb.save(filename)


if __name__ == '__main__':
    # get_gene_id('KLF2')
    # get_summary_info('10365')

    # result = get_summary_info(get_gene_id('ZNF805'))
    # pprint.pprint(result)

    # read_excel_gene_symbol(data_path)
    # main_process()

    print(get_summary_info('91'))
