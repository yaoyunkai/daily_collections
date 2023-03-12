"""



Create at 2023/3/12 11:08
"""
import glob
import json
import os

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

BASE_DIR = r'C:\Users\10524\Downloads'

HEADERS = [
    "id",
    "node_id",
    "name",
    "full_name",
    "private",
    "owner.login",
    "owner.id",
    "owner.html_url",
    "owner.type",
    "html_url",
    "description",
    "fork",
    "url",
    "languages_url",
    "commits_url",
    "git_commits_url",
    "ssh_url",
    "clone_url",
    "homepage",
    "stargazers_count",
    "watchers_count",
    "language",
    "forks_count",
    "archived",
    "disabled",
    "topics",  # list
    "updated_at",
    "created_at",
]


def read_repos_from_file():
    # for filename in os.listdir('data'):
    #     glob.glob(filename, )
    results = []

    for idx, filepath in enumerate(glob.glob('data/repos-*.json')):
        repos_data = json.load(open(filepath, mode='r'))
        for item in repos_data:
            one_repo_data = []
            for key in HEADERS:
                if '.' in key:
                    lhs, rhs = key.split('.')
                    _value = item.get(lhs, {}).get(rhs, '')

                elif key == 'topics':
                    _value = item.get(key, [])
                    if not _value:
                        _value = ''
                    else:
                        _value = ', '.join([str(i) for i in _value])
                else:
                    _value = item.get(key, '')
                one_repo_data.append(_value)
            # pprint.pprint(one_repo_data)
            results.append(one_repo_data)

    return results


def save_data_to_excel(data_list):
    """
    WorkBook
    WorkSheet
    Cell


    :return:
    """
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = 'Repository'
    # cell1 = ws['A1']
    # cell1 = ws['A2']

    for item in data_list:
        ws.append(item)

    file_path = os.path.join(BASE_DIR, 'repos.xlsx')
    wb.save(file_path)


if __name__ == '__main__':
    # read_repos_from_file()
    save_data_to_excel(read_repos_from_file())
